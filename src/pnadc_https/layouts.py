"""Read IBGE XLS/XLSX fixed-width dictionaries into normalized JSON."""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile

from .utils import atomic_json


@dataclass(slots=True)
class Variable:
    name: str
    start: int
    width: int
    end: int
    label: str = ""
    section: str = ""
    question: str = ""
    period: str = ""
    categories: dict[str, str] | str = field(default_factory=dict)
    storage_type: str = "string"


@dataclass(slots=True)
class Layout:
    variables: list[Variable]
    source: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": 1,
            "source": self.source,
            "variables": [asdict(variable) for variable in self.variables],
        }

    @classmethod
    def from_dict(cls, value: dict[str, Any]) -> "Layout":
        return cls(
            variables=[Variable(**item) for item in value["variables"]],
            source=dict(value.get("source") or {}),
        )


def _clean(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        result = re.sub(r"\s+", " ", value.replace('"', " ")).strip()
        try:
            numeric = float(result.replace(",", "."))
            return int(numeric) if numeric.is_integer() else numeric
        except ValueError:
            return result
    return value


def _rows_xls(data: bytes) -> list[list[object]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("xlrd is required for .xls dictionaries") from exc
    sheet = xlrd.open_workbook(file_contents=data).sheet_by_index(0)
    return [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]


def _rows_xlsx(data: bytes) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for .xlsx dictionaries") from exc
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _storage_type(width: int) -> str:
    if width == 15:
        return "float64"
    if width <= 2:
        return "int8"
    if width <= 4:
        return "int16"
    if width <= 9:
        return "int32"
    if width <= 18:
        return "int64"
    return "string"


def parse_rows(rows: Iterable[Iterable[object]], source: dict[str, str] | None = None) -> Layout:
    """Parse the layout convention used by PNADC dictionary workbooks."""
    variables: list[Variable] = []
    current: Variable | None = None
    section = ""
    for raw_row in rows:
        fields = [_clean(value) for value in raw_row]
        fields.extend([""] * max(0, 8 - len(fields)))
        if fields[0] and not fields[1] and not isinstance(fields[0], (int, float)):
            section = str(fields[0]).lower()
            continue
        if isinstance(fields[0], int) and isinstance(fields[1], int) and fields[2]:
            start, width = int(fields[0]), int(fields[1])
            categories: dict[str, str] | str = {}
            if fields[5] or fields[6]:
                category_hint = str(fields[5]).lower()
                categorical = isinstance(fields[5], (int, float)) or any(
                    token in category_hint for token in (" a ", "código", "valor", "130", "01-")
                )
                if categorical:
                    categories = {str(fields[5]): str(fields[6]).lower()}
                else:
                    categories = ", ".join(str(item).lower() for item in fields[5:7] if item)
            current = Variable(
                name=str(fields[2]).lower(),
                start=start,
                width=width,
                end=start + width - 1,
                question=str(fields[3]),
                label=str(fields[4]).lower() or str(fields[2]).lower(),
                section=section,
                period=str(fields[7]).lower(),
                categories=categories,
                storage_type=_storage_type(width),
            )
            variables.append(current)
            continue
        if current is not None and not fields[0] and not fields[1] and (fields[5] or fields[6]):
            if not isinstance(current.categories, dict):
                continue
            key = str(fields[5]) if fields[5] != "" else "vazio"
            current.categories[key] = str(fields[6]).lower()
    if not variables:
        raise ValueError("No fixed-width variables were found in the workbook")
    names = [variable.name for variable in variables]
    if len(names) != len(set(names)):
        raise ValueError("The layout contains duplicate variable names")
    return Layout(variables=variables, source=source or {})


def _select_dictionary_member(archive: ZipFile, member: str | None = None) -> str:
    if member:
        if member not in archive.namelist():
            raise ValueError(f"ZIP member not found: {member}")
        return member
    candidates = [
        name for name in archive.namelist()
        if name.lower().endswith((".xls", ".xlsx"))
        and any(token in Path(name).name.lower() for token in ("dicion", "input", "layout"))
    ]
    if not candidates:
        candidates = [name for name in archive.namelist() if name.lower().endswith((".xls", ".xlsx"))]
    if len(candidates) != 1:
        raise ValueError("Specify --member; the ZIP does not contain exactly one dictionary workbook")
    return candidates[0]


def load_layout(path: str | Path, member: str | None = None) -> Layout:
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix == ".json":
        with source_path.open("r", encoding="utf-8") as stream:
            return Layout.from_dict(json.load(stream))
    source = {"path": str(source_path.resolve())}
    if suffix == ".zip":
        with ZipFile(source_path) as archive:
            selected = _select_dictionary_member(archive, member)
            data = archive.read(selected)
        suffix = Path(selected).suffix.lower()
        source["member"] = selected
    else:
        data = source_path.read_bytes()
    if suffix == ".xls":
        rows = _rows_xls(data)
    elif suffix == ".xlsx":
        rows = _rows_xlsx(data)
    else:
        raise ValueError(f"Unsupported layout format: {suffix}")
    return parse_rows(rows, source)


def write_layout(layout: Layout, output: str | Path) -> None:
    atomic_json(Path(output), layout.to_dict())
